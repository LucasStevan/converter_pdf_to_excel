#pyinstaller --onefile --noconsole --icon="icone.ico" conversor.py

import os
import pandas as pd
import camelot
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook
from openpyxl.styles import Alignment

def extrair_dados_tabela_camelot(pdf_path):
    """Extrai dados de tabelas em PDFs usando camelot."""
    try:
        tabelas = camelot.read_pdf(pdf_path, pages="all", flavor="stream")
        if tabelas.n > 0:
            df_completo = pd.concat([tabela.df for tabela in tabelas], ignore_index=True)
            return df_completo
        else:
            return pd.DataFrame()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao processar o PDF: {e}")
        return pd.DataFrame()

def ajustar_colunas_worksheet(worksheet):
    """Configura as células para não invadir a próxima coluna."""
    for row in worksheet.iter_rows():  
        for cell in row:  
            if cell.value: 
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False, shrink_to_fit=True)

def salvar_dados_excel(lista_pdfs, output_path="resultado.xlsx"):
    """Salva os dados de cada PDF em uma aba no Excel."""
    if not os.path.exists(output_path):
        wb = Workbook()
        wb.save(output_path)

    total_pdfs = len(lista_pdfs)
    progresso["maximum"] = total_pdfs
    progresso["value"] = 0

    with pd.ExcelWriter(output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for i, pdf_path in enumerate(lista_pdfs, start=1):
            nome_aba = os.path.basename(pdf_path).replace(".pdf", "")[:31]
            df = extrair_dados_tabela_camelot(pdf_path)
            if not df.empty:
                df.to_excel(writer, sheet_name=nome_aba, index=False)

            progresso["value"] = i
            progresso_label.config(text=f"Convertendo {nome_aba}... ({i}/{total_pdfs})")
            root.update_idletasks()

    # Ajustar colunas após salvar os dados
    workbook = writer.book
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        ajustar_colunas_worksheet(worksheet)

    # Salvar o arquivo atualizado
    workbook.save(output_path)

    messagebox.showinfo("Sucesso", f"Arquivo salvo como {output_path}")
    progresso_label.config(text="Conversão concluída!")
    progresso["value"] = 0

def selecionar_pdfs():
    """Abre a janela para selecionar PDFs na pasta C:\."""
    arquivos = filedialog.askopenfilenames(initialdir="C:\\", filetypes=[("Arquivos PDF", "*.pdf")])

    if not arquivos:
        messagebox.showwarning("Nenhum arquivo", "Nenhum PDF foi selecionado.")
        return

    if len(arquivos) > 10:
        messagebox.showerror("Erro", "Selecione no máximo 10 arquivos PDF.")
        return

    salvar_dados_excel(arquivos)

root = tk.Tk()
root.title("Conversor PDF para Excel")
root.geometry("400x250")

# Carregar a imagem de plano de fundo
try:
    imagem_fundo = Image.open("yourlogo_.png")  # Substitua pelo nome correto do arquivo
    imagem_fundo = imagem_fundo.resize((400, 250))  # Redimensionar imagem para o tamanho da janela
    imagem_fundo_tk = ImageTk.PhotoImage(imagem_fundo)

    canvas = tk.Canvas(root, width=400, height=250)
    canvas.pack(fill="both", expand=True)
    canvas.create_image(0, 0, image=imagem_fundo_tk, anchor="nw")
except Exception as e:
    messagebox.showerror("Erro", f"Erro ao carregar o plano de fundo: {e}")
    canvas = tk.Canvas(root, width=400, height=250)
    canvas.pack(fill="both", expand=True)


botao_selecionar = tk.Button(root, text="Escolher PDF", command=selecionar_pdfs, font=("Arial", 12))
progresso = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
progresso_label = tk.Label(root, text="Aguardando seleção de PDFs...", font=("Arial", 10))


canvas.create_window(200, 50, window=botao_selecionar)  # Posição do botão
canvas.create_window(200, 100, window=progresso)        # Posição da barra de progresso
canvas.create_window(200, 140, window=progresso_label)  # Posição do label de progresso

root.mainloop()
